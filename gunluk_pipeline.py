"""
gunluk_pipeline.py
---------------------
Her gun otomatik calisacak ana betik. Sirasiyla:
1. TEFAS'tan bugunun fiyatlarini ceker
2. Fiyat gecmisini gunceller (yeni gunu ekler, en eskiyi atar -- 255 gunluk
   pencere korunur)
3. Donemsel getiri + Sharpe + MDD + Consistency + Otokorelasyon +
   EnBuyukGunOrani hesaplayip Z skorunu uretir
4. Tema kumelemeyi (TemaOlustur) calistirir
5. Sirala, dunku sirayi kaydet, docs/data.json'i yazar (dashboard'un
   okudugu dosya)

Bu betik Babur'un mudahalesine gerek kalmadan GitHub Actions tarafindan
her gun calistirilacak.
"""

import json
import datetime
from pathlib import Path

import numpy as np
from pytefas import Crawler, TefasAPIError, TefasRateLimitError

from puanlama_metrikleri import (
    donemsel_getiriler, sharpe, mdd, consistency, otokorelasyon,
    en_buyuk_gun_orani, puanlama_motoru, tema_olustur,
)

try:
    import openpyxl
except ImportError:
    openpyxl = None

KOK = Path(__file__).parent
GUN_SAYISI = 255

# --- FAZ 1 KAPSAM SINIRI ---
# Babur'un talebiyle: once sadece fiyat cekme + skor + siralama + tema +
# data.json (goruntu) kismi dogru calissin, o dogrulandiktan sonra emir
# motoru (AL/SAT islemleri) tekrar acilacak. False iken bekleyen_emirler.json
# HIC DOKUNULMADAN oldugu gibi kalir (ne islenir ne silinir) -- Faz 2'de
# True yapildiginda kaldigi yerden islenmeye devam eder.
EMIR_MOTORU_AKTIF = False


def bugunun_tarihi_veya_son_is_gunu() -> str:
    bugun = datetime.date.today()
    while bugun.weekday() >= 5:
        bugun = bugun - datetime.timedelta(days=1)
    return bugun.strftime("%Y-%m-%d")


def yeni_fon_gecmisini_cek(kod: str, gun_sayisi: int = GUN_SAYISI):
    """Yeni eklenen bir fon icin gecmis fiyat serisini TEFAS'tan ceker.
    Donus: (fon_adi, fiyatlar_listesi_yeniden_eskiye) ya da (None, None)."""
    bugun = datetime.date.today()
    baslangic = bugun - datetime.timedelta(days=int(gun_sayisi * 1.5))  # hafta sonlari icin pay
    tefas = Crawler()
    df = tefas.fetch(baslangic.strftime("%Y-%m-%d"), bugun.strftime("%Y-%m-%d"),
                      columns="info", kind="YAT", fund_code=kod)
    kayitlar = json.loads(df.to_json(orient="records"))
    if not kayitlar:
        return None, None
    # tarihe gore yeniden eskiye sirala
    kayitlar.sort(key=lambda k: k.get("date", ""), reverse=True)
    fiyatlar = [k.get("price", 0.0) for k in kayitlar][:gun_sayisi]
    ad = kayitlar[0].get("fund_name")
    if len(fiyatlar) < gun_sayisi:
        # yetersiz gecmis (yeni kurulmus fon olabilir) -- ilk degerle doldur
        doldurma = fiyatlar[-1] if fiyatlar else 0.0
        fiyatlar = fiyatlar + [doldurma] * (gun_sayisi - len(fiyatlar))
    return ad, fiyatlar


def yukle(dosya):
    with open(KOK / dosya, encoding="utf-8") as f:
        return json.load(f)


def tarihte_fiyat(tarihler, fiyatlar, hedef_tarih):
    """tarihler (en yeniden eskiye siralı) icinde hedef_tarih'e esit ya da
    ondan once gelen ilk (en yakin) fiyati dondurur. Bulunamazsa None."""
    for i, t in enumerate(tarihler):
        if t and t <= hedef_tarih:
            return fiyatlar[i]
    return None


def ay_yil_baslangici(tarih_str):
    """'YYYY-MM-DD' -> (ay_baslangici, yil_baslangici)ayni formatta."""
    yil, ay = tarih_str[:4], tarih_str[5:7]
    return f"{yil}-{ay}-01", f"{yil}-01-01"


def is_gunu_ekle(tarih_str, gun_sayisi):
    """tarih_str ('YYYY-MM-DD') uzerine gun_sayisi kadar IS GUNU (hafta
    sonu haric) ekler. Valor hesaplamasi icin (resmi tatiller dahil
    edilmiyor, yaklasik bir tahmindir)."""
    tarih = datetime.date.fromisoformat(tarih_str)
    kalan = int(gun_sayisi)
    while kalan > 0:
        tarih += datetime.timedelta(days=1)
        if tarih.weekday() < 5:  # 0-4 = Pazartesi-Cuma
            kalan -= 1
    return tarih.strftime("%Y-%m-%d")


def eksik_gunu_excelden_doldur(fon_listesi: dict, fiyat_gecmisi: dict) -> None:
    """Babur'un TEFAS'tan elle okuyup 'eksik_gun_manuel.xlsx' sablonuna
    girdigi bir gunu price_history.json'a isler. Kod bilmedigi icin JSON
    yerine Excel kullaniyor. Islem basariyla bitince (ya da tarih zaten
    mevcutsa) sablondaki Tarih ve Fiyat hucreleri otomatik bosaltilir --
    ayni dosya bir sonraki eksik gun icin tekrar kullanilabilir."""
    yol = KOK / "eksik_gun_manuel.xlsx"
    if not yol.is_file():
        return
    if openpyxl is None:
        print("UYARI: openpyxl kurulu degil, eksik_gun_manuel.xlsx islenemedi.")
        return

    wb = openpyxl.load_workbook(yol)
    ws = wb["Eksik Gun"] if "Eksik Gun" in wb.sheetnames else wb.active
    hedef_tarih = str(ws["B9"].value or "").strip()
    if not hedef_tarih:
        return  # sablon bos, yapilacak bir sey yok

    HDR_ROW = 11
    try:
        datetime.date.fromisoformat(hedef_tarih)
        gecerli_tarih = True
    except ValueError:
        print(f"UYARI: eksik_gun_manuel.xlsx'teki tarih ('{hedef_tarih}') "
              f"YYYY-AA-GG formatinda degil, islem yapilmadi ama sablon temizlendi.")
        gecerli_tarih = False

    if gecerli_tarih:
        if hedef_tarih in fiyat_gecmisi["tarihler"]:
            print(f"eksik_gun_manuel.xlsx: {hedef_tarih} zaten fiyat gecmisinde var, atlaniyor.")
        else:
            fiyatlar_girilen = {}
            r = HDR_ROW + 1
            while ws.cell(row=r, column=1).value:
                kod = str(ws.cell(row=r, column=1).value).strip()
                deger = ws.cell(row=r, column=3).value
                if deger not in (None, "") and kod in fon_listesi:
                    try:
                        fiyatlar_girilen[kod] = float(deger)
                    except (TypeError, ValueError):
                        print(f"  UYARI: {kod} icin gecersiz fiyat ('{deger}'), atlandi.")
                r += 1

            # tarihler azalan (yeniden eskiye) sirali -- hedef_tarih'in
            # gitmesi gereken indeksi bul (ayni mantik eksik_gun_doldur.json ile)
            ekleme_idx = len(fiyat_gecmisi["tarihler"])
            for i, t in enumerate(fiyat_gecmisi["tarihler"]):
                if t < hedef_tarih:
                    ekleme_idx = i
                    break
            fiyat_gecmisi["tarihler"].insert(ekleme_idx, hedef_tarih)
            fiyat_gecmisi["tarihler"] = fiyat_gecmisi["tarihler"][:GUN_SAYISI]
            for kod in fon_listesi:
                seri = fiyat_gecmisi["fiyatlar"].get(kod, [0.0] * GUN_SAYISI)
                komsu = seri[ekleme_idx] if ekleme_idx < len(seri) else (
                    seri[ekleme_idx - 1] if ekleme_idx > 0 and ekleme_idx - 1 < len(seri) else 0.0)
                doldurulacak = fiyatlar_girilen.get(kod, komsu)
                seri.insert(ekleme_idx, doldurulacak)
                fiyat_gecmisi["fiyatlar"][kod] = seri[:GUN_SAYISI]
            kaydet(fiyat_gecmisi, "price_history.json")
            eksik_sayisi = len(fon_listesi) - len(fiyatlar_girilen)
            print(f"eksik_gun_manuel.xlsx: {hedef_tarih} eklendi -- {len(fiyatlar_girilen)} fon "
                  f"Excel'den girilen gercek fiyatla, {eksik_sayisi} fon komsu gunun fiyatiyla "
                  f"(indeks {ekleme_idx}).")

    # sablonu temizle -- Tarih ve tum Fiyat hucreleri bosalsin, bir sonraki
    # eksik gun icin ayni dosya tekrar kullanilabilsin
    ws["B9"] = None
    r = HDR_ROW + 1
    while ws.cell(row=r, column=1).value:
        ws.cell(row=r, column=3).value = None
        r += 1
    wb.save(yol)


def kaydet(veri, dosya):
    with open(KOK / dosya, "w", encoding="utf-8") as f:
        json.dump(veri, f, ensure_ascii=False, indent=1)


def eksik_gunleri_bul(tarihler: list, yoksayilanlar: list = (), en_fazla: int = 10) -> list:
    """price_history.json'daki tarih dizisinde (en yeniden en eskiye dogru
    sirali) hafta ici oldugu halde atlanmis gunleri tespit eder. Sadece
    hafta sonlarini eler; resmi tatiller nedeniyle TEFAS'in zaten
    yayinlamadigi gunler burada da 'eksik' olarak gorunebilir -- bu
    yanlis alarm degil, sadece TEFAS o gun veri yayinlamadiysa gercekten
    doldurulacak bir sey olmadigi anlamina gelir. Kullanici arayuzden
    boyle bir gunu 'yoksay' dediginde tarih eksik_gun_yoksay.json'a
    eklenir ve bundan sonra kalici olarak listeye hic girmez. En guncel
    en_fazla kaydi doner (cok eski bosluklarla listeyi sismemek icin)."""
    if len(tarihler) < 2:
        return []
    yoksay_set = set(yoksayilanlar)
    eksikler = []
    for i in range(len(tarihler) - 1):
        sonraki = datetime.date.fromisoformat(tarihler[i])   # daha yeni
        onceki = datetime.date.fromisoformat(tarihler[i + 1])  # daha eski
        gun = onceki + datetime.timedelta(days=1)
        while gun < sonraki:
            if gun.weekday() < 5 and gun.isoformat() not in yoksay_set:  # 0=Pzt ... 4=Cuma, hafta ici
                eksikler.append(gun.isoformat())
            gun += datetime.timedelta(days=1)
    eksikler.sort(reverse=True)
    return eksikler[:en_fazla]


def bugunku_fiyatlari_cek(fon_kodlari: set, tarih: str):
    """TEFAS'tan o gune ait tum fon fiyatlarini ceker, sadece takip
    listemizdeki fonlari dondurur: (fiyatlar_dict, tefas_tarihi).
    tefas_tarihi, TEFAS'in kendi dondurdugu 'date' alanindan okunur --
    yerel tarih tahminine guvenilmez (hafta sonu/tatil/henuz yayinlanmamis
    veri durumlarinda yanlis olabilir)."""
    tefas = Crawler()
    df = tefas.fetch(tarih, columns="info", kind="YAT")
    kayitlar = json.loads(df.to_json(orient="records"))
    sonuc = {}
    tefas_tarihi = None
    for k in kayitlar:
        kod = k.get("fund_code")
        if kod in fon_kodlari:
            sonuc[kod] = k.get("price")
        if tefas_tarihi is None and k.get("date"):
            tefas_tarihi = str(k["date"])[:10]
    return sonuc, tefas_tarihi


def main():
    fon_listesi = yukle("fon_listesi.json")
    sabitler = yukle("sabitler.json")
    fiyat_gecmisi = yukle("price_history.json")
    try:
        onceki_rank = yukle("onceki_rank.json")
    except FileNotFoundError:
        onceki_rank = {}
    try:
        portfoy = yukle("portfoy.json")  # {kod: {adet, alis_tarihi?}} -- elde tutulan fonlar
    except FileNotFoundError:
        portfoy = {}
    try:
        bekleyen_emirler = yukle("bekleyen_emirler.json")
    except FileNotFoundError:
        bekleyen_emirler = []
    try:
        bekleyen_valorler = yukle("bekleyen_valorler.json")
    except FileNotFoundError:
        bekleyen_valorler = []
    try:
        eklenecek_fonlar = yukle("eklenecek_fonlar.json")
    except FileNotFoundError:
        eklenecek_fonlar = []
    try:
        eksik_gun_yoksay = yukle("eksik_gun_yoksay.json")  # kullanicinin "bu gun tatil, eksik degil" dedigi tarihler
    except FileNotFoundError:
        eksik_gun_yoksay = []

    # --- yeni eklenen fonlarin 1 yillik gecmisini cek ---
    if eklenecek_fonlar:
        print(f"{len(eklenecek_fonlar)} yeni fon icin gecmis veri cekiliyor...")
        kalan_eklenecek = []
        for istek in eklenecek_fonlar:
            kod = istek.get("kod")
            valor = istek.get("valor")
            if kod in fon_listesi:
                print(f"  {kod} zaten listede, atlandi.")
                continue
            try:
                ad, fiyatlar = yeni_fon_gecmisini_cek(kod)
            except (TefasRateLimitError, TefasAPIError) as e:
                print(f"  HATA: {kod} icin veri cekilemedi ({e}), tekrar denenecek.")
                kalan_eklenecek.append(istek)
                continue
            if not ad:
                print(f"  UYARI: {kod} icin TEFAS'ta veri bulunamadi, kod hatali olabilir.")
                continue
            fon_listesi[kod] = {"ad": ad, "valor": valor}
            fiyat_gecmisi["fiyatlar"][kod] = fiyatlar
            print(f"  {kod} ({ad}) eklendi, {len(fiyatlar)} gunluk gecmisle.")
        eklenecek_fonlar = kalan_eklenecek
        kaydet(eklenecek_fonlar, "eklenecek_fonlar.json")
        kaydet(fon_listesi, "fon_listesi.json")
        kaydet(fiyat_gecmisi, "price_history.json")

    # --- eksik/atlanmis bir gunu doldur (tek seferlik talimat dosyasi) ---
    try:
        eksik_gun = yukle("eksik_gun_doldur.json")
    except FileNotFoundError:
        eksik_gun = {}
    if eksik_gun.get("tarih"):
        hedef_tarih = eksik_gun["tarih"]
        if hedef_tarih in fiyat_gecmisi["tarihler"]:
            print(f"{hedef_tarih} zaten fiyat gecmisinde var, atlaniyor.")
        else:
            print(f"Eksik gun dolduruluyor: {hedef_tarih}")
            eksik_fiyatlar, dogrulanan_tarih = bugunku_fiyatlari_cek(set(fon_listesi.keys()), hedef_tarih)
            if dogrulanan_tarih != hedef_tarih:
                print(f"  UYARI: TEFAS bu tarih icin '{dogrulanan_tarih}' dondurdu "
                      f"(istenen: {hedef_tarih}) -- islem yapilmadi, tarihi kontrol edin.")
            else:
                # tarihler azalan (yeniden eskiye) sirali -- hedef_tarih'in
                # gitmesi gereken indeksi bul (ilk index'in tarihi <
                # hedef_tarih olan yer)
                ekleme_idx = len(fiyat_gecmisi["tarihler"])
                for i, t in enumerate(fiyat_gecmisi["tarihler"]):
                    if t < hedef_tarih:
                        ekleme_idx = i
                        break
                fiyat_gecmisi["tarihler"].insert(ekleme_idx, hedef_tarih)
                fiyat_gecmisi["tarihler"] = fiyat_gecmisi["tarihler"][:GUN_SAYISI]
                for kod in fon_listesi:
                    seri = fiyat_gecmisi["fiyatlar"].get(kod, [0.0] * GUN_SAYISI)
                    komsu = seri[ekleme_idx] if ekleme_idx < len(seri) else (
                        seri[ekleme_idx - 1] if ekleme_idx > 0 and ekleme_idx - 1 < len(seri) else 0.0)
                    doldurulacak = eksik_fiyatlar.get(kod, komsu)
                    seri.insert(ekleme_idx, doldurulacak)
                    fiyat_gecmisi["fiyatlar"][kod] = seri[:GUN_SAYISI]
                kaydet(fiyat_gecmisi, "price_history.json")
                print(f"  {hedef_tarih} eklendi ({len(eksik_fiyatlar)} fon icin gercek fiyatla), "
                      f"indeks {ekleme_idx}.")
        kaydet({}, "eksik_gun_doldur.json")  # tek seferlik -- tuketildi

    # --- eksik/atlanmis bir gunu Excel sablonundan doldur (Babur kod bilmedigi icin) ---
    eksik_gunu_excelden_doldur(fon_listesi, fiyat_gecmisi)

    tarih_tahmin = bugunun_tarihi_veya_son_is_gunu()
    print(f"Tahmini tarih: {tarih_tahmin} (TEFAS'in kendi tarihiyle dogrulanacak)")

    print("TEFAS'tan bugunun fiyatlari cekiliyor...")
    try:
        bugun_fiyat, tefas_tarihi = bugunku_fiyatlari_cek(set(fon_listesi.keys()), tarih_tahmin)
    except (TefasRateLimitError, TefasAPIError) as e:
        print(f"HATA: TEFAS'tan veri cekilemedi: {e}")
        return

    if not tefas_tarihi:
        print("HATA: TEFAS'tan tarih bilgisi alinamadi, islem durduruldu.")
        return

    son_bilinen_tarih = fiyat_gecmisi["tarihler"][0] if fiyat_gecmisi["tarihler"] else None
    yeni_gun_var_mi = (tefas_tarihi != son_bilinen_tarih)
    tarih = tefas_tarihi
    print(f"TEFAS'in gercek veri tarihi: {tarih} (elimizdeki en yeni: {son_bilinen_tarih}) "
          f"-> {'YENI GUN' if yeni_gun_var_mi else 'henuz yeni veri yok, pencere kaydirilmayacak'}")

    eksik = [k for k in fon_listesi if k not in bugun_fiyat and k not in ("END", "BBR")]
    if eksik:
        print(f"UYARI: {len(eksik)} fon icin bugun fiyat gelmedi (islem "
              f"gormemis olabilir): {eksik[:10]}{'...' if len(eksik)>10 else ''}")

    # "onceki_gun_indeksi": END hesaplarken ve genel getiri karsilastirmalarinda
    # "dunku fiyat" olarak hangi sutuna bakacagimizi belirler.
    # - Yeni bir gun ise: index 0 henuz DUNKU fiyati tutuyor (bugunku fiyati
    #   az sonra basa ekleyecegiz), yani onceki gun = index 0.
    # - Ayni gunun tekrar calistirilmasiysa (sabah birkac kez guncelleme --
    #   Excel'deki eski calisma seklinizin ayni): index 0 zaten BUGUNUN
    #   (kismi/eksik) fiyatini tutuyor, gercek onceki gun index 1'de.
    onceki_gun_indeksi = 0 if yeni_gun_var_mi else 1

    # --- END sentetik endeksi (Endeks.bas / EndeksHesapla ile ayni mantik) ---
    # Her calistirmada yeniden hesaplanir (sabah birkac kez guncellemede de
    # dahil) -- boylece eksik fonlar sonradan gelince END de dogru guncellenir.
    haric = {"END", "BBR"}
    getiriler = []
    for kod in fon_listesi:
        if kod in haric:
            continue
        onceki_seri = fiyat_gecmisi["fiyatlar"].get(kod, [0.0])
        dun_fiyat = onceki_seri[onceki_gun_indeksi] if len(onceki_seri) > onceki_gun_indeksi else 0.0
        bugun_fiyat_kod = bugun_fiyat.get(kod)
        if dun_fiyat and bugun_fiyat_kod:
            getiriler.append((bugun_fiyat_kod - dun_fiyat) / dun_fiyat * 100)
    if getiriler:
        ort_getiri = sum(getiriler) / len(getiriler)
        end_onceki_seri = fiyat_gecmisi["fiyatlar"].get("END", [0.0])
        end_dun = end_onceki_seri[onceki_gun_indeksi] if len(end_onceki_seri) > onceki_gun_indeksi else 0.0
        if not end_dun or end_dun <= 0:
            end_dun = 100.0  # ilk calistirma tabani
        bugun_fiyat["END"] = end_dun * (1 + ort_getiri / 100)
        print(f"END guncellendi: ortalama getiri %{ort_getiri:.4f} ({len(getiriler)} fon), "
              f"yeni deger {bugun_fiyat['END']:.4f}")

    # --- fiyat gecmisi guncellemesi ---
    if yeni_gun_var_mi:
        # gercekten yeni bir gun: pencereyi kaydir, bugunun (o an elde
        # olan) fiyatlarini basa ekle
        fiyat_gecmisi["tarihler"].insert(0, tarih)
        fiyat_gecmisi["tarihler"] = fiyat_gecmisi["tarihler"][:GUN_SAYISI]
        for kod in fon_listesi:
            eski_seri = fiyat_gecmisi["fiyatlar"].get(kod, [0.0] * GUN_SAYISI)
            yeni_fiyat = bugun_fiyat.get(kod, eski_seri[0] if eski_seri else 0.0)
            eski_seri = [yeni_fiyat] + eski_seri
            fiyat_gecmisi["fiyatlar"][kod] = eski_seri[:GUN_SAYISI]
        print(f"Pencere kaydirildi, yeni gun eklendi ({len(bugun_fiyat)} fon fiyati ile).")
    else:
        # ayni gunun tekrar calistirilmasi: pencereyi KAYDIRMA, sadece
        # bugunun (index 0) sutununu -- yeni gelen fiyatlarla -- guncelle.
        # Boylece sabah eksik olan bir fon, sonraki calistirmada
        # tamamlanabiliyor (Excel'deki "tum fiyatlar gelene kadar
        # tekrar tekrar guncelle" aliskanliginizin karsiligi).
        guncellenen_sayisi = 0
        for kod, yeni_fiyat in bugun_fiyat.items():
            seri = fiyat_gecmisi["fiyatlar"].get(kod)
            if seri:
                if seri[0] != yeni_fiyat:
                    guncellenen_sayisi += 1
                seri[0] = yeni_fiyat
        print(f"Pencere kaydirilmadi -- bugunun sutunu {guncellenen_sayisi} fonda "
              f"tazelendi, skorlar yeniden hesaplanacak.")

    # --- GUNLUK KAZANC BIRIKTIRICILERI (Ay/Yil sutunlari) ---
    # KRITIK SIRALAMA: bu blok, bekleyen AL/SAT emirleri islenmeden ONCE
    # calisir. Boylece bugun SAT edilecek bir fonun o gunku getirisi,
    # portfoyden silinmeden once Aylik/Yillik biriktiriciye eklenmis olur.
    # adet burada portfoy'un DUNDEN devreden hali (henuz bugunku emirler
    # uygulanmadi) -- bugun AL edilen bir fon icin adet=0 cikar, yani o
    # fonun bugunku kazanci otomatik 0 olur (ayrica ozel durum gerekmez).
    try:
        fon_kazanc = yukle("fon_kazanc.json")
    except FileNotFoundError:
        fon_kazanc = {}
    bugun_ay_kk, bugun_yil_kk = tarih[:7], tarih[:4]
    gunluk_kazanc = {}
    for kod in fon_listesi:
        pozisyon = portfoy.get(kod)
        adet_dun = pozisyon.get("adet", 0) if isinstance(pozisyon, dict) else (pozisyon or 0)
        seri = fiyat_gecmisi["fiyatlar"].get(kod, [0.0, 0.0])
        f_bugun = seri[0] if len(seri) > 0 else 0.0
        f_dun = seri[1] if len(seri) > 1 else f_bugun
        kazanc = round(adet_dun * (f_bugun - f_dun), 2) if adet_dun else 0.0
        gunluk_kazanc[kod] = kazanc

        girdi = fon_kazanc.get(kod, {})
        if girdi.get("ay") != bugun_ay_kk:
            girdi = {**girdi, "ay": bugun_ay_kk, "ay_deger": 0.0}
        if girdi.get("yil") != bugun_yil_kk:
            girdi = {**girdi, "yil": bugun_yil_kk, "yil_deger": 0.0}
        girdi["ay_deger"] = round(girdi.get("ay_deger", 0.0) + kazanc, 2)
        girdi["yil_deger"] = round(girdi.get("yil_deger", 0.0) + kazanc, 2)
        fon_kazanc[kod] = girdi
    kaydet(fon_kazanc, "fon_kazanc.json")

    # --- bekleyen AL/SAT emirlerini isle (Rutin'in 212/213 mantiginin karsiligi) ---
    if not EMIR_MOTORU_AKTIF:
        if bekleyen_emirler:
            print(f"FAZ 1: emir motoru kapali, {len(bekleyen_emirler)} bekleyen emire "
                  f"DOKUNULMADI (bekleyen_emirler.json degismedi).")
    elif bekleyen_emirler:
        print(f"{len(bekleyen_emirler)} bekleyen emir isleniyor...")
        kalan_emirler = []
        for emir in bekleyen_emirler:
            kod = emir.get("kod")
            tip = emir.get("tip")
            fiyat_bugun_kod = bugun_fiyat.get(kod)
            if fiyat_bugun_kod is None or fiyat_bugun_kod <= 0:
                print(f"  UYARI: {kod} icin bugun fiyat yok, emir ertelendi.")
                kalan_emirler.append(emir)
                continue
            if tip == "AL":
                tutar = emir.get("tutar", 0)
                adet = int(tutar / fiyat_bugun_kod)
                portfoy[kod] = {"adet": adet, "alis_tarihi": tarih}
                print(f"  AL: {kod} {adet} adet ({tutar} TL / {fiyat_bugun_kod})")
            elif tip == "SAT":
                # Non-PPF fonlar sadece BLOK satilir -- emirde adet
                # gonderilse bile yok sayilir, mevcut pozisyonun tamami satilir.
                mevcut = portfoy.get(kod)
                mevcut_adet = mevcut.get("adet", 0) if isinstance(mevcut, dict) else (mevcut or 0)
                satilacak_adet = mevcut_adet
                satis_tutari = round(satilacak_adet * fiyat_bugun_kod, 2)
                valor_gun = fon_listesi.get(kod, {}).get("valor") or 0
                hesaba_gecis = is_gunu_ekle(tarih, valor_gun)
                bekleyen_valorler.append({
                    "kod": kod, "tutar": satis_tutari,
                    "satis_tarihi": tarih, "hesaba_gecis_tarihi": hesaba_gecis,
                })
                if kod in portfoy:
                    del portfoy[kod]
                print(f"  SAT (blok): {kod} pozisyonu tamamen kapatildi "
                      f"({satilacak_adet} adet, {satis_tutari} TL "
                      f"{hesaba_gecis} tarihinde hesaba gecer)")
        bekleyen_emirler = kalan_emirler  # sadece ertelenenler kalir
        kaydet(bekleyen_emirler, "bekleyen_emirler.json")
        kaydet(portfoy, "portfoy.json")

    # --- suresi gecmis (hesaba gecmis) valor hatirlaticilarini temizle ---
    bekleyen_valorler = [v for v in bekleyen_valorler if v["hesaba_gecis_tarihi"] > tarih]
    kaydet(bekleyen_valorler, "bekleyen_valorler.json")

    kodlar = list(fon_listesi.keys())
    fiyat_full = np.array([fiyat_gecmisi["fiyatlar"][k] for k in kodlar])
    n = sabitler["periyod"]
    fiyat_np1 = fiyat_full[:, : n + 1]
    fiyat_n = fiyat_full[:, :n]

    print("Skorlar hesaplaniyor...")
    k_sigmoid = sabitler["k"]
    dk = sabitler["donemsel_katsayilar"]
    Z = np.zeros(len(kodlar))
    donemsel = donemsel_getiriler(fiyat_full)
    donem_map = {"gunluk": "gunluk", "haftalik": "haftalik", "aylik": "aylik",
                 "uc_aylik": "uc_aylik", "alti_aylik": "alti_aylik", "yillik": "yillik"}
    for ad, kat_anahtar in donem_map.items():
        Z += puanlama_motoru(donemsel[ad], yon=1, katsayi=dk[kat_anahtar], k=k_sigmoid, winsor=True)

    topk6 = sum(dk.values())
    oran = sabitler["risk_metrik_oranlari"]
    Z += puanlama_motoru(sharpe(fiyat_np1, sabitler["risksiz_yillik"]), yon=1,
                          katsayi=topk6 * oran["sharpe"], k=k_sigmoid, winsor=True)
    Z += puanlama_motoru(mdd(fiyat_n), yon=-1,
                          katsayi=topk6 * oran["mdd"], k=k_sigmoid, winsor=False)
    Z += puanlama_motoru(consistency(fiyat_np1), yon=1,
                          katsayi=topk6 * oran["consistency"], k=k_sigmoid, winsor=True)
    Z += puanlama_motoru(otokorelasyon(fiyat_np1), yon=-1,
                          katsayi=topk6 * oran["otokorelasyon"], k=k_sigmoid, winsor=True)
    Z += puanlama_motoru(en_buyuk_gun_orani(fiyat_np1), yon=-1,
                          katsayi=topk6 * oran["ebg"], k=k_sigmoid, winsor=False)

    print("Tema kumeleme calisiyor...")
    valorler = np.array([fon_listesi[k].get("valor") for k in kodlar], dtype=object)
    fiyat_91 = fiyat_full[:, :91]
    tema_sonuc = dict(tema_olustur(fiyat_91, kodlar, valorler))

    # --- Is gunu sayilari (TarihKontrol'un G216/G217 karsiligi) ---
    # tarih formati YYYY-MM-DD; ay/yil karsilastirmasi bugunku (en yeni) tarihe gore
    bugun_ay, bugun_yil = tarih[:7], tarih[:4]
    ay_sayisi = sum(1 for t in fiyat_gecmisi["tarihler"] if t and t[:7] == bugun_ay)
    yil_sayisi = sum(1 for t in fiyat_gecmisi["tarihler"] if t and t[:4] == bugun_yil)

    sira = np.argsort(-Z)

    # --- Tema catisma kurali: rank1 ve rank2 ayni temaysa dogrudan yer degistir ---
    # Sıradaki ilk FARKLI temali fon rank2 olur; eski rank2 fonu, o fonun
    # eski sirasina gecer (duz yer degistirme, aradakiler kaymaz).
    tema_swap_oldu = False
    if len(sira) >= 3:
        def _tema_grubu(kod):
            t = tema_sonuc.get(kod)
            return t.split(".")[0] if t else None

        top1_kod, top2_kod = kodlar[sira[0]], kodlar[sira[1]]
        tema1 = _tema_grubu(top1_kod)
        if tema1 is not None and tema1 == _tema_grubu(top2_kod):
            for j in range(2, len(sira)):
                aday_kod = kodlar[sira[j]]
                if _tema_grubu(aday_kod) != tema1:
                    sira = sira.copy()
                    sira[1], sira[j] = sira[j], sira[1]
                    tema_swap_oldu = True
                    print(f"Tema catismasi: {top1_kod} ve {top2_kod} ayni temada ({tema1}). "
                          f"{aday_kod} yeni rank2 oldu, {top2_kod} sira {j+1}'e tasindi.")
                    break

    yeni_rank = {}
    dashboard_funds = []
    for sirano, idx in enumerate(sira, start=1):
        kod = kodlar[idx]
        yeni_rank[kod] = sirano
        dunku = onceki_rank.get(kod)
        pozisyon = portfoy.get(kod)
        adet = pozisyon.get("adet", 0) if isinstance(pozisyon, dict) else (pozisyon or 0)
        alis_tarihi = pozisyon.get("alis_tarihi") if isinstance(pozisyon, dict) else None
        fiyat_bugun = float(fiyat_full[idx, 0])
        fiyat_dun = float(fiyat_full[idx, 1]) if fiyat_full.shape[1] > 1 else fiyat_bugun
        guncel = round(adet * fiyat_bugun, 2)
        # Bugun ALINAN bir fon icin gosterilecek gunluk getiri her zaman 0
        # (o gunku fiyat hareketini sahiplenmedi) -- diger tum gunlerde
        # gunluk_kazanc[kod] (emirler islenmeden ONCE, dunden devreden adetle
        # hesaplanmisti) kullanilir; SAT ile bugun kapanan pozisyonlar icin de
        # bu deger zaten dogru hesaplanip biriktiriciye eklenmis durumda.
        if alis_tarihi == tarih:
            gunluk_tl = 0.0
        else:
            gunluk_tl = gunluk_kazanc.get(kod, round(adet * (fiyat_bugun - fiyat_dun), 2))

        # Aylik/Yillik kar-zarar: artik kalici biriktiriciden okunuyor
        # (fon_kazanc.json) -- gunluk kazancin fon bazinda ay/yil basinda
        # sifirlanarak biriktirilmis hali. PPF/blok ayrimi ya da maliyet
        # bazina ihtiyac yok; fon portfoyden cikmis olsa bile o ay/yil
        # icindeki son degeri donmus halde gorunmeye devam eder.
        kazanc_girdi = fon_kazanc.get(kod, {})
        ay_kar = kazanc_girdi.get("ay_deger", 0.0)
        yil_kar = kazanc_girdi.get("yil_deger", 0.0)

        dashboard_funds.append({
            "kod": kod,
            "ad": fon_listesi[kod].get("ad"),
            "val": fon_listesi[kod].get("valor"),
            "rnk": sirano,
            "dunRnk": dunku,
            "isBold": bool(dunku is not None and sirano <= dunku),
            "gun": round(float(donemsel["gunluk"][idx]), 4),
            "haf": round(float(donemsel["haftalik"][idx]), 4),
            "ay": round(float(donemsel["aylik"][idx]), 4),
            "ay3": round(float(donemsel["uc_aylik"][idx]), 4),
            "ay6": round(float(donemsel["alti_aylik"][idx]), 4),
            "yil": round(float(donemsel["yillik"][idx]), 4),
            "tema": tema_sonuc.get(kod),
            "adet": adet,             # elde tutulan pay adedi
            "guncel": guncel,         # elde tutulan fonun bugunku TL degeri (adet x fiyat)
            "gunlukTL": gunluk_tl,    # bugunku TL bazli kar/zarar
            "ayKarZarar": ay_kar,     # ay basindan beri biriken TL kazanc (fon_kazanc.json)
            "yilKarZarar": yil_kar,   # yil basindan beri biriken TL kazanc (fon_kazanc.json)
        })


    portfolio = [f for f in dashboard_funds if f["guncel"] and f["guncel"] > 0]
    toplam_guncel_deger = round(sum(f["guncel"] for f in portfolio), 2)

    # --- TOPLAM Gunluk/Aylik/Yillik kar-zarar: fon_kazanc bazli ---
    # Sutun basliklarindaki toplam, artik o sutundaki tum fonlarin (elde
    # olsun olmasin) degerlerinin toplami. Bugun SAT ile kapanan bir
    # pozisyonun kazanci gunluk_kazanc/fon_kazanc icinde zaten var, bu
    # yuzden "portfolio" (sadece su an elde tutulanlar) uzerinden DEGIL,
    # tum fon_listesi uzerinden toplaniyor.
    totals = {
        "daily": round(sum(gunluk_kazanc.values()), 2),
        "monthly": round(sum(v.get("ay_deger", 0.0) for v in fon_kazanc.values()), 2),
        "yearly": round(sum(v.get("yil_deger", 0.0) for v in fon_kazanc.values()), 2),
    }

    data_json = {
        "funds": dashboard_funds,
        "portfolio": portfolio,
        "totals": totals,
        "timestamp": int(datetime.datetime.now().timestamp() * 1000),
        "tableDate": tarih,
        "isGunuSayilari": {"aylik": ay_sayisi, "yillik": yil_sayisi},
        "bekleyenValorler": bekleyen_valorler,
        "eksikFonlar": eksik,
        "temaSwapOldu": tema_swap_oldu,
        "eksikGunler": eksik_gunleri_bul(fiyat_gecmisi["tarihler"], eksik_gun_yoksay),
    }

    kaydet(fiyat_gecmisi, "price_history.json")
    kaydet(yeni_rank, "onceki_rank.json")
    kaydet(data_json, "docs/data.json")
    print(f"Tamamlandi: {len(dashboard_funds)} fon icin docs/data.json yazildi.")


if __name__ == "__main__":
    main()
